"""
Excel 导出模块
将一级/二级提炼结果回写到 DataFrame，并生成带合并单元格的 Excel。
"""

import json
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


def export_to_excel(
    data: pd.DataFrame,
    level1_file: str,
    level2_file: str,
    output_path: str = "output_kh.xlsx",
):
    """
    将一级、二级 Know-How 结果写入 Excel，二级结果按批次合并单元格。

    Parameters
    ----------
    data : 完整的原始数据 DataFrame
    level1_file : 一级提炼结果 JSON 路径
    level2_file : 二级压缩结果 JSON 路径
    output_path : 输出 Excel 路径
    """
    with open(level1_file, "r", encoding="utf-8") as f:
        level1 = json.load(f)

    with open(level2_file, "r", encoding="utf-8") as f:
        level2 = json.load(f)

    for i in range(len(level1)):
        l1 = level1[str(i)]
        data.loc[i, "know-how"] = l1["Know_How"]
        data.loc[i, "Logic_Diagnosis"] = l1["Logic_Diagnosis"]

    row_to_l2_key = {}
    for k, v in level2.items():
        for row_idx in v["source_indices"]:
            row_to_l2_key[row_idx] = k

    data["二级提炼KH"] = None
    for row_idx in data.index:
        l2_key = row_to_l2_key.get(row_idx)
        if l2_key is not None:
            data.at[row_idx, "二级提炼KH"] = level2[l2_key]["Final_Know_How"]

    print(f"二级 KH 覆盖行数: {data['二级提炼KH'].notna().sum()} / {len(data)}")

    data.to_excel(output_path, index=False, engine="openpyxl")

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_l2 = header.index("二级提炼KH") + 1

    index_list = list(data.index)
    label_to_xl = {label: pos + 2 for pos, label in enumerate(index_list)}

    l2_batch_xl_rows = {}
    for label, xl_row in label_to_xl.items():
        l2_key = row_to_l2_key.get(label)
        if l2_key is not None:
            l2_batch_xl_rows.setdefault(l2_key, []).append(xl_row)

    for l2_key, xl_rows in sorted(l2_batch_xl_rows.items()):
        if len(xl_rows) <= 1:
            continue
        xl_rows.sort()
        ws.merge_cells(
            start_row=xl_rows[0], start_column=col_l2,
            end_row=xl_rows[-1], end_column=col_l2,
        )
        ws.cell(row=xl_rows[0], column=col_l2).alignment = Alignment(
            wrap_text=True, vertical="top",
        )

    wb.save(output_path)
    n = len(data)
    print(f"已保存至 {output_path}（共 {n} 行数据，"
          f"二级批次 {len(l2_batch_xl_rows)} 组合并）")
    return output_path
